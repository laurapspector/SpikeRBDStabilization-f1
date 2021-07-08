from copy import copy
import sys

import numpy as np

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, GradientFill
)
from openpyxl.styles.fills import Stop
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

NORMAL_COLOR = 'FFFFFF'
MISSING_COLOR = 'BBBBBB'
MISSING_FILL = PatternFill(start_color=MISSING_COLOR,
                           end_color=MISSING_COLOR,
                           fill_type='solid')

LABEL_COL_DIMS = dict(A=12, B=12, C=8, D=3)
COL_DIM = 16 / 6

AA_DISPLAY_ORDER = list('*FWYPMILVAGCSTNQDEHKR')

AA_CAT_LABEL_START_ROW = 3
AA_CAT_LABEL_END_COL = 3
AA_CAT_LABEL_MERGED_CELLS = {
    'hydrophobic' : CellRange('A4:A12'),
    'hydrophilic' : CellRange('A13:A23'),
    'aromatic' : CellRange('B4:B6'),
    'non-polar\naliphatic' : CellRange('B7:B13'),
    'polar uncharged' : CellRange('B14:B18'),
    'negatively\ncharged' : CellRange('B19:B20'),
    'positively\ncharged' : CellRange('B21:B23'),
    'small' : CellRange('C12:C15'),
    'START' : CellRange('C8:C8'),
    'STOP' : CellRange('C3:C3')
}
AA_LABEL_START_ROW = AA_CAT_LABEL_START_ROW
AA_LABEL_COL = AA_CAT_LABEL_END_COL + 1
POS_LABEL_ROW = AA_CAT_LABEL_START_ROW - 2
WT_AA_LABEL_ROW = POS_LABEL_ROW + 1
DATA_START_ROW = POS_LABEL_ROW + 2
DATA_START_COL = AA_LABEL_COL + 1

def add_heatmap_aa_labels(ws):
    for col, width in LABEL_COL_DIMS.items():
        ws.column_dimensions[col].width = width

    for name, cell_range in AA_CAT_LABEL_MERGED_CELLS.items():
        start_col, start_row, end_col, end_row = cell_range.bounds
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=end_row, end_column=end_col)
        ws.cell(start_row, start_col).alignment = Alignment(horizontal='center',
                                                            vertical='center',
                                                            wrapText=True)
        ws.cell(start_row, start_col).font = Font(name='Helvetica')
        ws.cell(start_row, start_col).value = name

        full_range = CellRange(min_row=start_row, min_col=start_col,
                               max_row=end_row, max_col=AA_CAT_LABEL_END_COL)

        for row, col in full_range.cells:
            border = copy(ws.cell(row, col).border)
            border.left = Side()
            border.top = Side()
            border.bottom = Side()
            ws.cell(row, col).border = border
        for side in ['left', 'top', 'bottom']:
            for row, col in getattr(full_range, side):
                border = copy(ws.cell(row, col).border)
                setattr(border, side, Side(border_style='thin'))
                ws.cell(row, col).border = border
        # for row, col in full_range.left:
        #     border = copy(ws.cell(row, col).border)
        #     border.left = Side(border_style='thin')
        #     ws.cell(row, col).border = border
        # for row, col in full_range.top:
        #     border = copy(ws.cell(row, col).border)
        #     border.top = Side(border_style='thin')
        #     ws.cell(row, col).border = border
        # for row, col in full_range.bottom:
        #     border = copy(ws.cell(row, col).border)
        #     border.bottom = Side(border_style='thin')
        #     ws.cell(row, col).border = border

    for row, aa in enumerate(AA_DISPLAY_ORDER, AA_LABEL_START_ROW):
        ws.cell(row, AA_LABEL_COL).alignment = Alignment(horizontal='center')
        ws.cell(row, AA_LABEL_COL).font = Font(name='Helvetica')
        ws.cell(row, AA_LABEL_COL).value = aa


def add_heatmap_data(ws, wt_seq, positions, datasets, significance,
                     pos_label_colors, hit_color):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    ws.row_dimensions[POS_LABEL_ROW].height = 24
    for col, pos in enumerate(positions, DATA_START_COL):
        wt_aa = wt_seq[pos]
        label_font = Font(name='Helvetica',
                          color=pos_label_colors.get(pos, '000000'))
        ws.column_dimensions[get_column_letter(col)].width = COL_DIM
        ws.cell(POS_LABEL_ROW, col).value = pos
        ws.cell(POS_LABEL_ROW, col).font = label_font
        ws.cell(POS_LABEL_ROW, col).alignment = Alignment(horizontal='center',
                                                          vertical='center',
                                                          textRotation=75)
        ws.cell(WT_AA_LABEL_ROW, col).value = wt_aa
        ws.cell(WT_AA_LABEL_ROW, col).font = label_font
        ws.cell(WT_AA_LABEL_ROW, col).alignment = Alignment(horizontal='center',
                                                            vertical='center')
        for row, aa in enumerate(AA_DISPLAY_ORDER, DATA_START_ROW):
            ws.cell(row, col).font = Font(name='Helvetica')
            ws.cell(row, col).alignment = Alignment(horizontal='center',
                                                    vertical='center')
            ws.cell(row, col).number_format = ';;;'
            ws.cell(row, col).border = thin_border
            colors = []
            ER = None
            for data in datasets:
                d = data.loc[data['variant'] == f'{wt_aa}{pos}{aa}']
                if len(d) == 0:
                    colors.append(MISSING_COLOR)
                elif len(d) == 1:
                    ER = d['ER'].to_numpy()[0]
                    pval = d['pval'].to_numpy()[0]
                    if pval <= significance:
                        colors.append(hit_color)
                    else:
                        colors.append(NORMAL_COLOR)
                else:
                    raise ValueError('Multiple ERs for mutation.')
            if len(datasets) == 0:
                pass
            elif len(datasets) == 1:
                if ER is not None: ws.cell(row, col).value = ER
                ws.cell(row, col).fill = PatternFill(fill_type='solid',
                                                     fgColor=colors[0])
            else:
                delta = 1e-8
                stop_positions = [i / len(colors) for i in range(len(colors)+1)]
                stops = []
                for i, color in enumerate(colors):
                    stops.append(Stop(color, stop_positions[i] + delta*(i>0)))
                    stops.append(Stop(color, stop_positions[i+1]))
                ws.cell(row, col).fill = GradientFill(stop=stops, degree=45)


def write_heatmap(filename, wt_seq, positions, data, title, significance,
                  pos_label_colors={}, hit_color='1F78B4'):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.sheet_view.showGridLines = False
    add_heatmap_aa_labels(ws)
    add_heatmap_data(ws, wt_seq, positions, data, significance,
                     pos_label_colors, hit_color)
    wb.save(filename)
